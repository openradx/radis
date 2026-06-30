"""Tests for ``export_collection`` (radis/collections/utils/exporters.py).

The view-level test only asserts the HTTP MIME type; here we actually open the
produced workbook with openpyxl and verify the header row plus the cell values
of collected reports (document fields, formatted dates, modalities).
"""

from datetime import UTC, date, datetime

import pytest
from adit_radis_shared.accounts.factories import UserFactory
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from radis.collections.factories import CollectionFactory
from radis.collections.utils.exporters import export_collection
from radis.reports.factories import LanguageFactory, ReportFactory

EXPECTED_HEADER = [
    "PACS",
    "Patient ID",
    "Patient Birth Date",
    "Study Date",
    "Study Description",
    "Study Instance UID",
    "Accession Number",
    "Modalities",
    "Content",
]


def _load_sheet(collection) -> Worksheet:
    file = export_collection(collection)
    file.seek(0)
    wb = load_workbook(file)
    ws = wb.active
    assert ws is not None
    return ws


@pytest.mark.django_db
def test_export_header_row():
    collection = CollectionFactory.create(owner=UserFactory.create())
    ws = _load_sheet(collection)

    header = [cell.value for cell in ws[1]]
    assert header == EXPECTED_HEADER


@pytest.mark.django_db
def test_export_report_cell_values():
    language = LanguageFactory.create(code="en")
    report = ReportFactory.create(
        language=language,
        pacs_name="Test PACS",
        patient_id="PAT-12345",
        patient_birth_date=date(1980, 1, 15),
        study_datetime=datetime(2024, 3, 15, 10, 30, tzinfo=UTC),
        study_description="CT Thorax",
        study_instance_uid="1.2.3.4.5",
        accession_number="ACC-9999",
        body="Findings: no acute abnormality.",
        modalities=["CT", "PT"],
    )
    collection = CollectionFactory.create(owner=UserFactory.create())
    collection.reports.add(report)

    ws = _load_sheet(collection)

    # Exactly one data row beneath the header.
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    row = rows[0]
    by_col = dict(zip(EXPECTED_HEADER, row))

    assert by_col["PACS"] == "Test PACS"
    assert by_col["Patient ID"] == "PAT-12345"
    # Dates are formatted via SHORT_DATE_FORMAT == "m/d/Y".
    assert by_col["Patient Birth Date"] == "01/15/1980"
    assert by_col["Study Date"] == "03/15/2024"
    assert by_col["Study Description"] == "CT Thorax"
    assert by_col["Study Instance UID"] == "1.2.3.4.5"
    assert by_col["Accession Number"] == "ACC-9999"
    # Modalities are joined with ", " in insertion order.
    modalities = by_col["Modalities"]
    assert isinstance(modalities, str)
    assert set(modalities.split(", ")) == {"CT", "PT"}
    assert by_col["Content"] == "Findings: no acute abnormality."


@pytest.mark.django_db
def test_export_multiple_reports_all_present():
    language = LanguageFactory.create(code="en")
    report1 = ReportFactory.create(language=language, patient_id="PAT-AAA")
    report2 = ReportFactory.create(language=language, patient_id="PAT-BBB")
    collection = CollectionFactory.create(owner=UserFactory.create())
    collection.reports.add(report1, report2)

    ws = _load_sheet(collection)

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    pid_col = EXPECTED_HEADER.index("Patient ID")
    patient_ids = {str(row[pid_col]) for row in rows}
    assert patient_ids == {"PAT-AAA", "PAT-BBB"}


@pytest.mark.django_db
def test_export_empty_collection_has_header_only():
    collection = CollectionFactory.create(owner=UserFactory.create())
    ws = _load_sheet(collection)

    header = [cell.value for cell in ws[1]]
    assert header == EXPECTED_HEADER
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows == []
